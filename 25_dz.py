import tkinter as tk
from openpyxl import load_workbook
import os


def count_rows(path):
    total_rows = 0
    for filename in os.listdir(path):
        if filename.endswith(".xlsx"):
            workbook = load_workbook(os.path.join(path, filename))
            for sheet in workbook.worksheets:
                total_rows += sheet.max_row
    return total_rows


def main():
    root = tk.Tk()
    root.title("Подсчет строк в Excel")
    root.geometry("400x200")
    result_label = tk.Label(root, text="Общее количество строк:")
    result_label.pack()

    count_button = tk.Button(root, text="Подсчитать",
                             command=lambda: update_result())
    count_button.pack()

    def update_result():
        path = "c:\Users\hawci\OneDrive\Рабочий стол\nurda\25_dz.py"
        total_rows = count_rows(path)
        result_label.config(text=f"Общее количество строк: {total_rows}")

    root.mainloop()


if __name__ == "__main__":
    main()